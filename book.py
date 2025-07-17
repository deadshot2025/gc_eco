import json
import pandas as pd
from openpyxl import load_workbook
import pytesseract
from PIL import Image
import cv2
import os
import re
from datetime import datetime

# Constants
TEMPLATE_PATH = "data/GC_ECO_Template.xlsx"
RATES_PATH = "data/vat_cit_usc_rates.json"
TAGMAP_PATH = "tag_map.md"


class GCEcoBook:
    def __init__(self):
        # Load the Excel workbook from TEMPLATE_PATH using openpyxl
        self.workbook = load_workbook(TEMPLATE_PATH)
        
        # Load JSON metadata from RATES_PATH into self.meta
        with open(RATES_PATH, 'r') as f:
            self.meta = json.load(f)
        
        # Call self._write_meta() to write metadata to Excel
        self._write_meta()
        
        # Load tag mapping from TAGMAP_PATH using pandas
        # Skip first row, use columns 1,2, handle markdown table format with regex separator
        tag_df = pd.read_csv(TAGMAP_PATH, sep=r"\s*\|\s*", engine='python', 
                            skiprows=1, usecols=[1, 2], names=['tag', 'target_sheet'])
        
        # Filter out the markdown separator row (contains dashes)
        tag_df = tag_df[~tag_df['tag'].str.contains('-', na=False)]
        
        # Create self.tag_map dictionary from tag->target_sheet mapping
        self.tag_map = dict(zip(tag_df['tag'], tag_df['target_sheet']))
    
    def _write_meta(self):
        # Get the GC_META worksheet
        meta_sheet = self.workbook['GC_META']
        
        # Write JSON-serialized self.meta to cell A1
        meta_sheet['A1'] = json.dumps(self.meta)
    
    def save(self):
        # Save the workbook to TEMPLATE_PATH
        self.workbook.save(TEMPLATE_PATH)
    
    def load_raw(self, df_raw: pd.DataFrame):
        # Get the RAW worksheet
        raw_sheet = self.workbook['RAW']
        
        # Iterate through df_raw rows and append each row as a list to the worksheet
        for _, row in df_raw.iterrows():
            raw_sheet.append(row.tolist())
        
        # Call save()
        self.save()
    
    def route_raw(self):
        # Read RAW worksheet data into a pandas DataFrame
        raw_sheet = self.workbook['RAW']
        
        # Extract all data from the worksheet
        data = []
        for row in raw_sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data.append(row)
        
        if not data:
            return  # No data to process
        
        # Extract headers from first row, set as column names
        headers = data[0]
        df_raw = pd.DataFrame(data[1:], columns=headers)
        
        # For each data row, check if any tag from self.tag_map appears in the Description column (case-insensitive)
        for _, row in df_raw.iterrows():
            description = str(row.get('Description', '')).lower()
            
            # Check if any tag appears in the description
            for tag, target_sheet in self.tag_map.items():
                if tag.lower() in description:
                    # If match found, call _append_row() with the target sheet and row data
                    self._append_row(target_sheet, row)
                    break  # Only route to first matching tag
        
        # Call save()
        self.save()
    
    def _append_row(self, sheet: str, row: pd.Series):
        # Get the specified worksheet
        target_sheet = self.workbook[sheet]
        
        # Append the row as a list to the worksheet
        target_sheet.append(row.tolist())
    
    def _preprocess_image(self, image_path: str) -> str:
        """
        Preprocess image for better OCR accuracy.
        
        Args:
            image_path (str): Path to the input image
            
        Returns:
            str: Path to the preprocessed image
            
        Raises:
            FileNotFoundError: If image file doesn't exist
            ValueError: If image format is not supported
            Exception: For other image processing errors
        """
        # Validate file exists
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image file not found: {image_path}")
        
        # Validate supported formats
        supported_formats = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        file_ext = os.path.splitext(image_path)[1].lower()
        if file_ext not in supported_formats:
            raise ValueError(f"Unsupported image format: {file_ext}. Supported formats: {supported_formats}")
        
        try:
            # Load image using OpenCV
            image = cv2.imread(image_path)
            if image is None:
                raise ValueError(f"Could not load image: {image_path}. File may be corrupted.")
            
            # Convert to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Apply Gaussian blur to reduce noise
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            
            # Apply threshold to get binary image (black text on white background)
            _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Create output path for preprocessed image
            base_name = os.path.splitext(os.path.basename(image_path))[0]
            output_dir = os.path.dirname(image_path)
            preprocessed_path = os.path.join(output_dir, f"{base_name}_preprocessed.png")
            
            # Save preprocessed image
            cv2.imwrite(preprocessed_path, thresh)
            
            return preprocessed_path
            
        except Exception as e:
            raise Exception(f"Error preprocessing image {image_path}: {str(e)}")
    
    def _extract_text_from_image(self, image_path: str) -> str:
        """
        Extract text from image using OCR.
        
        Args:
            image_path (str): Path to the image file
            
        Returns:
            str: Extracted text from the image
            
        Raises:
            FileNotFoundError: If image file doesn't exist
            ValueError: If image format is not supported or corrupted
            Exception: For OCR processing errors
        """
        # Validate file exists and is readable
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image file not found: {image_path}")
        
        # Validate supported formats
        supported_formats = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.pdf'}
        file_ext = os.path.splitext(image_path)[1].lower()
        if file_ext not in supported_formats:
            raise ValueError(f"Unsupported image format: {file_ext}. Supported formats: {supported_formats}")
        
        try:
            # Handle PDF files differently
            if file_ext == '.pdf':
                # For PDF files, use PIL to convert first page to image
                # import fitz  # PyMuPDF - would need to be added to requirements if PDF support needed
                # For now, raise an exception as PDF support requires additional dependency
                raise ValueError("PDF support not implemented. Please convert PDF to image format first.")
            
            # Preprocess image for better OCR accuracy
            preprocessed_path = self._preprocess_image(image_path)
            
            try:
                # Load preprocessed image using PIL
                pil_image = Image.open(preprocessed_path)
                
                # Extract text using pytesseract
                # Use custom config for better accuracy
                custom_config = r'--oem 3 --psm 6'
                text = pytesseract.image_to_string(pil_image, config=custom_config)
                
                # Clean up preprocessed image
                if os.path.exists(preprocessed_path) and preprocessed_path != image_path:
                    os.remove(preprocessed_path)
                
                return text.strip()
                
            except Exception as ocr_error:
                # Clean up preprocessed image on error
                if os.path.exists(preprocessed_path) and preprocessed_path != image_path:
                    os.remove(preprocessed_path)
                raise ocr_error
                
        except pytesseract.TesseractNotFoundError:
            raise Exception(
                "Tesseract OCR not found. Please install Tesseract OCR:\n"
                "- Ubuntu/Debian: sudo apt-get install tesseract-ocr\n"
                "- macOS: brew install tesseract\n"
                "- Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki"
            )
        except (FileNotFoundError, ValueError):
            # Re-raise these specific exceptions as-is
            raise
        except Exception as e:
            raise Exception(f"Error extracting text from image {image_path}: {str(e)}")
    
    def _extract_receipt_data(self, ocr_text: str) -> dict:
        """
        Extract structured data from OCR text of a receipt.
        
        Args:
            ocr_text (str): Raw OCR text from receipt image
            
        Returns:
            dict: Structured transaction data with keys: Date, Description, Amount, VATRate, Tags
        """
        # Initialize result with defaults
        result = {
            "Date": None,
            "Description": "",
            "Amount": 0.0,
            "VATRate": 23,  # Default Irish VAT rate
            "Tags": ""
        }
        
        # Date parsing regex patterns
        date_patterns = [
            r'(\d{4})[\/\-\.](\d{1,2})[\/\-\.](\d{1,2})',  # YYYY/MM/DD or YYYY-MM-DD
            r'(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{4})',  # DD/MM/YYYY
            r'(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2})',  # DD/MM/YY
        ]
        
        # Amount parsing regex patterns
        amount_patterns = [
            r'TOTAL[:\s]*€?\s*(\d+[,\.]\d{2})',  # TOTAL: €XX.XX
            r'€\s*(\d+[,\.]\d{2})',  # €XX.XX
            r'(\d+[,\.]\d{2})\s*€',  # XX.XX €
            r'(\d+[,\.]\d{2})\s*EUR',  # XX.XX EUR
        ]
        
        # Extract date
        for i, pattern in enumerate(date_patterns):
            match = re.search(pattern, ocr_text)
            if match:
                groups = match.groups()
                try:
                    if i == 0:  # YYYY/MM/DD format
                        year, month, day = groups
                    elif i == 1:  # DD/MM/YYYY format
                        day, month, year = groups
                    else:  # DD/MM/YY format
                        day, month, year = groups
                        year = f"20{year}" if int(year) < 50 else f"19{year}"
                    
                    # Validate date
                    date_obj = datetime.strptime(f"{year}-{month:0>2}-{day:0>2}", "%Y-%m-%d")
                    # Check if date is reasonable (not too far in past/future)
                    current_year = datetime.now().year
                    if 2000 <= date_obj.year <= current_year + 1:
                        result["Date"] = date_obj.strftime("%Y-%m-%d")
                        break
                except (ValueError, TypeError):
                    continue
        
        # Extract amount
        for pattern in amount_patterns:
            match = re.search(pattern, ocr_text, re.IGNORECASE)
            if match:
                amount_str = match.group(1)
                try:
                    # Handle both comma and dot as decimal separator
                    amount_str = amount_str.replace(',', '.')
                    amount = float(amount_str)
                    if amount > 0:  # Validate positive amount
                        result["Amount"] = amount
                        break
                except (ValueError, TypeError):
                    continue
        
        # Extract merchant/business name (usually first few lines)
        lines = [line.strip() for line in ocr_text.split('\n') if line.strip()]
        merchant_name = ""
        if lines:
            # Take first non-empty line as merchant name
            merchant_name = lines[0]
            # Clean up common OCR errors
            merchant_name = re.sub(r'[^\w\s]', ' ', merchant_name)
            merchant_name = ' '.join(merchant_name.split())
        
        # Create description
        if merchant_name:
            result["Description"] = f"{merchant_name} - receipt"
        else:
            result["Description"] = "Receipt transaction"
        
        # Determine VAT rate from text if available
        vat_patterns = [
            r'VAT[:\s]*\((\d+)%\)',  # VAT (13%)
            r'VAT[:\s]*(\d+)%',      # VAT: 13%
            r'BTW[:\s]*(\d+)%',      # Dutch VAT
            r'(\d+)%\s*VAT',         # 13% VAT
            r'\((\d+)%\)',           # (13%) - generic percentage in parentheses
        ]
        
        for pattern in vat_patterns:
            match = re.search(pattern, ocr_text, re.IGNORECASE)
            if match:
                try:
                    vat_rate = int(match.group(1))
                    if 0 <= vat_rate <= 30:  # Reasonable VAT rate range
                        result["VATRate"] = vat_rate
                        break
                except (ValueError, TypeError):
                    continue
        
        # Basic categorization based on merchant name
        merchant_lower = merchant_name.lower()
        if any(keyword in merchant_lower for keyword in ['grocery', 'supermarket', 'market', 'food']):
            result["Tags"] = "grocery"
        elif any(keyword in merchant_lower for keyword in ['restaurant', 'cafe', 'coffee', 'bar', 'pub']):
            result["Tags"] = "restaurant"
        elif any(keyword in merchant_lower for keyword in ['office', 'supplies', 'stationery']):
            result["Tags"] = "office"
        elif any(keyword in merchant_lower for keyword in ['fuel', 'gas', 'petrol', 'station']):
            result["Tags"] = "fuel"
        else:
            result["Tags"] = "expense"
        
        return result
    
    def process_receipt_image(self, image_path: str) -> dict:
        """
        Process a single receipt image to extract structured transaction data.
        
        Args:
            image_path (str): Path to the receipt image file
            
        Returns:
            dict: Structured transaction data
            
        Raises:
            FileNotFoundError: If image file doesn't exist
            ValueError: If image format is not supported
            Exception: For OCR or parsing errors
        """
        try:
            # Extract text from image using OCR
            ocr_text = self._extract_text_from_image(image_path)
            
            # Parse structured data from OCR text
            receipt_data = self._extract_receipt_data(ocr_text)
            
            return receipt_data
            
        except (FileNotFoundError, ValueError):
            # Re-raise these specific exceptions as-is
            raise
        except Exception as e:
            raise Exception(f"Error processing receipt image {image_path}: {str(e)}")
    
    def process_receipts_folder(self, folder_path: str = "data/receipts") -> pd.DataFrame:
        """
        Process all receipt images in a folder and return structured data as DataFrame.
        
        Args:
            folder_path (str): Path to folder containing receipt images
            
        Returns:
            pd.DataFrame: DataFrame with columns matching RAW sheet format
        """
        # Supported image extensions
        supported_extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        
        # Check if folder exists
        if not os.path.exists(folder_path):
            print(f"Warning: Folder {folder_path} does not exist")
            return pd.DataFrame(columns=["Date", "Description", "Amount", "VATRate", "Tags"])
        
        # Collect all results
        results = []
        processed_count = 0
        error_count = 0
        
        # Iterate through all files in the folder
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            
            # Skip directories and non-image files
            if not os.path.isfile(file_path):
                continue
            
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext not in supported_extensions:
                continue
            
            try:
                # Process the receipt image
                receipt_data = self.process_receipt_image(file_path)
                results.append(receipt_data)
                processed_count += 1
                print(f"Processed: {filename}")
                
            except Exception as e:
                error_count += 1
                print(f"Warning: Failed to process {filename}: {str(e)}")
                continue
        
        print(f"Batch processing complete: {processed_count} processed, {error_count} errors")
        
        # Create DataFrame from results
        if results:
            df = pd.DataFrame(results)
            # Ensure columns are in the correct order
            df = df[["Date", "Description", "Amount", "VATRate", "Tags"]]
            return df
        else:
            return pd.DataFrame(columns=["Date", "Description", "Amount", "VATRate", "Tags"])


if __name__ == "__main__":
    # Create GCEcoBook instance
    book = GCEcoBook()
    
    # Example usage for receipt processing
    print("=== GC Eco Book - Receipt Processing Demo ===\n")
    
    # Process single receipt
    print("1. Processing single receipt:")
    try:
        receipt_data = book.process_receipt_image("data/receipts/test_receipt.png")
        print(f"   Date: {receipt_data['Date']}")
        print(f"   Description: {receipt_data['Description']}")
        print(f"   Amount: €{receipt_data['Amount']}")
        print(f"   VAT Rate: {receipt_data['VATRate']}%")
        print(f"   Tags: {receipt_data['Tags']}")
    except Exception as e:
        print(f"   Error: {e}")
    
    print("\n2. Batch processing receipts folder:")
    try:
        df = book.process_receipts_folder("data/receipts")
        print(f"   Processed {len(df)} receipts")
        if len(df) > 0:
            print("   DataFrame:")
            print(df.to_string(index=False))
    except Exception as e:
        print(f"   Error: {e}")
    
    print("\n3. Loading and routing receipt data:")
    try:
        if len(df) > 0:
            book.load_raw(df)
            book.route_raw()
            print("   Receipt data loaded and routed successfully!")
    except Exception as e:
        print(f"   Error: {e}")
    
    # Example usage for loading CSV and routing (commented)
    # df = pd.read_csv('data/receipts/sample_receipts.csv')
    # book.load_raw(df)
    # book.route_raw()
    
    # Call save()
    book.save()
    print("\n4. Workbook saved successfully!")
