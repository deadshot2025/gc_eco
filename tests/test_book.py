import pandas as pd
from openpyxl import load_workbook
import pytest
from book import GCEcoBook


def test_meta_loaded():
    """Test that metadata is properly loaded from JSON."""
    # Create GCEcoBook instance
    eco = GCEcoBook()
    
    # Assert that "VAT" is in eco.meta
    assert "VAT" in eco.meta
    
    # Test that metadata is properly loaded from JSON
    assert eco.meta["VAT"]["standard"] == 23
    assert eco.meta["VAT"]["reduced"] == 8
    assert eco.meta["VAT"]["zero"] == 0
    assert eco.meta["CIT"]["standard"] == 19
    assert eco.meta["CIT"]["small_business"] == 9
    assert eco.meta["USC"]["rate"] == 4.5


def test_routing():
    """Test that transaction routing works correctly."""
    # Create GCEcoBook instance
    eco = GCEcoBook()
    
    # Create test DataFrame with sample data
    test_data = {
        "Date": "2025-07-01",
        "Description": "office expense",
        "Amount": 100,
        "VATRate": 23,
        "Tags": ""
    }
    df = pd.DataFrame([test_data])
    
    # Call eco.load_raw(df) to load the test data
    eco.load_raw(df)
    
    # Call eco.route_raw() to route the data
    eco.route_raw()
    
    # Load the Excel workbook and verify that EXPENSES sheet contains the routed data
    workbook = load_workbook("data/GC_ECO_Template.xlsx")
    expenses_sheet = workbook["EXPENSES"]
    
    # Check for Amount=100 in any row
    found_amount = False
    for row in expenses_sheet.iter_rows(values_only=True):
        if row and len(row) > 2 and row[2] == 100:  # Amount is in column 3 (index 2)
            found_amount = True
            break
    
    assert found_amount, "Expected Amount=100 not found in EXPENSES sheet"